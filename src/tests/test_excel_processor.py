"""
Unit tests for the Excel processor module.
"""

import pytest
import tempfile
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.rag.excel_processor import ExcelProcessor, ExcelWorksheetInfo, CellRange
from src.rag.document_processor import (
    DocumentType,
    ProcessingStatus,
    DocumentProcessingError,
    ExtractedImage,
    ExtractedTable
)


class TestExcelProcessor:
    """Test cases for ExcelProcessor class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.config = {
            'max_file_size_mb': 10,
            'image_processing': True,
            'table_extraction': True,
            'min_table_rows': 2,
            'min_table_cols': 2,
            'max_empty_rows': 5,
            'max_empty_cols': 5
        }
        self.processor = ExcelProcessor(self.config)
    
    def test_initialization(self):
        """Test Excel processor initialization."""
        assert self.processor.extract_images is True
        assert self.processor.extract_tables is True
        assert self.processor.min_table_rows == 2
        assert self.processor.min_table_cols == 2
        assert self.processor.max_empty_rows == 5
        assert self.processor.max_empty_cols == 5
    
    def test_get_supported_extensions(self):
        """Test supported extensions."""
        extensions = self.processor._get_supported_extensions()
        assert '.xlsx' in extensions
        assert '.xls' in extensions
        assert '.xlsm' in extensions
    
    def test_can_process(self):
        """Test file type detection."""
        assert self.processor.can_process('test.xlsx')
        assert self.processor.can_process('document.XLSX')  # Case insensitive
        assert self.processor.can_process('test.xls')
        assert self.processor.can_process('test.xlsm')
        assert not self.processor.can_process('test.txt')
        assert not self.processor.can_process('test.pdf')
    
    def test_count_non_empty_cells(self):
        """Test counting non-empty cells."""
        # Create a mock worksheet
        mock_worksheet = Mock()
        mock_worksheet.iter_rows.return_value = [
            ('Value1', 'Value2', None),
            (None, 'Value3', ''),
            ('Value4', None, 'Value5')
        ]
        
        count = self.processor._count_non_empty_cells(mock_worksheet)
        assert count == 4  # 'Value1', 'Value2', 'Value3', 'Value4', 'Value5' but empty string doesn't count
    
    def test_find_data_blocks(self):
        """Test finding data blocks in worksheet."""
        # Create a mock worksheet with data pattern
        mock_worksheet = Mock()
        mock_worksheet.max_row = 6
        mock_worksheet.max_column = 3
        mock_worksheet.iter_rows.return_value = [
            ('Header1', 'Header2', 'Header3'),  # Row 1
            ('Data1', 'Data2', 'Data3'),        # Row 2
            ('Data4', 'Data5', 'Data6'),        # Row 3
            (None, None, None),                 # Row 4 - empty
            (None, None, None),                 # Row 5 - empty
            ('NewHeader1', 'NewHeader2', None)   # Row 6 - new block
        ]
        
        data_blocks = self.processor._find_data_blocks(mock_worksheet)
        
        # Should find at least one data block
        assert len(data_blocks) >= 1
        if data_blocks:
            first_block = data_blocks[0]
            assert len(first_block) >= 2  # At least header + one data row
    
    def test_extract_worksheet_content(self):
        """Test extracting content from worksheet."""
        # Create a mock worksheet
        mock_worksheet = Mock()
        mock_worksheet.max_row = 3
        mock_worksheet.max_column = 2
        mock_worksheet.iter_rows.return_value = [
            ('Name', 'Age'),
            ('John', 25),
            ('Jane', 30)
        ]
        
        content = self.processor._extract_worksheet_content(mock_worksheet, 'Sheet1')
        
        assert 'Name\tAge' in content
        assert 'John\t25' in content
        assert 'Jane\t30' in content
    
    def test_extract_worksheet_content_empty(self):
        """Test extracting content from empty worksheet."""
        mock_worksheet = Mock()
        mock_worksheet.max_row = None
        
        content = self.processor._extract_worksheet_content(mock_worksheet, 'Sheet1')
        assert content == ""
    
    def test_extract_worksheet_content_with_none_values(self):
        """Test extracting content with None values."""
        mock_worksheet = Mock()
        mock_worksheet.max_row = 2
        mock_worksheet.max_column = 3
        mock_worksheet.iter_rows.return_value = [
            ('Name', None, 'City'),
            ('John', 25, None)
        ]
        
        content = self.processor._extract_worksheet_content(mock_worksheet, 'Sheet1')
        
        assert 'Name\t\tCity' in content
        assert 'John\t25\t' in content
    
    @patch('openpyxl.load_workbook')
    def test_extract_metadata(self, mock_load_workbook):
        """Test metadata extraction."""
        # Create mock workbook with properties
        mock_workbook = Mock()
        mock_props = Mock()
        mock_props.title = 'Test Workbook'
        mock_props.creator = 'Test Author'
        mock_props.description = 'Test Description'
        mock_props.created = Mock()
        mock_props.created.isoformat.return_value = '2023-01-01T00:00:00'
        mock_props.modified = Mock()
        mock_props.modified.isoformat.return_value = '2023-01-02T00:00:00'
        
        mock_workbook.properties = mock_props
        mock_workbook.sheetnames = ['Sheet1', 'Sheet2']
        mock_workbook.active.title = 'Sheet1'
        
        # Create temporary file for file info
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(b'fake excel content')
            tmp_path = tmp.name
        
        try:
            metadata = self.processor._extract_metadata(mock_workbook, tmp_path)
            
            assert metadata['title'] == 'Test Workbook'
            assert metadata['creator'] == 'Test Author'
            assert metadata['description'] == 'Test Description'
            assert metadata['created'] == '2023-01-01T00:00:00'
            assert metadata['modified'] == '2023-01-02T00:00:00'
            assert metadata['worksheet_count'] == 2
            assert metadata['active_sheet'] == 'Sheet1'
            assert 'file_size' in metadata
            
        finally:
            Path(tmp_path).unlink()
    
    def test_extract_image_data(self):
        """Test extracting image data from openpyxl Image."""
        # Create mock image
        mock_image = Mock()
        mock_image._data.return_value = b'fake_image_data'
        mock_image.format = 'PNG'
        
        image_data = self.processor._extract_image_data(mock_image)
        
        assert image_data is not None
        assert image_data['content'] == b'fake_image_data'
        assert image_data['format'] == 'PNG'
    
    def test_extract_image_data_no_data(self):
        """Test extracting image data when no data available."""
        mock_image = Mock()
        mock_image._data = None
        mock_image.ref = None
        
        image_data = self.processor._extract_image_data(mock_image)
        assert image_data is None
    
    def test_extract_defined_table(self):
        """Test extracting a defined Excel table."""
        # Create mock table
        mock_table = Mock()
        mock_table.ref = 'A1:C3'
        mock_table.name = 'TestTable'
        
        # Create mock worksheet
        mock_worksheet = Mock()
        mock_worksheet.__getitem__.return_value = [
            [Mock(value='Header1'), Mock(value='Header2'), Mock(value='Header3')],
            [Mock(value='Data1'), Mock(value='Data2'), Mock(value='Data3')],
            [Mock(value='Data4'), Mock(value='Data5'), Mock(value='Data6')]
        ]
        mock_table.parent = mock_worksheet
        
        table = self.processor._extract_defined_table(mock_table, 'Sheet1', 0)
        
        assert table is not None
        assert table.table_id == 'Sheet1_table0'
        assert table.headers == ['Header1', 'Header2', 'Header3']
        assert len(table.rows) == 2
        assert table.rows[0] == ['Data1', 'Data2', 'Data3']
        assert table.worksheet_name == 'Sheet1'
        assert table.cell_range == 'A1:C3'
    
    def test_process_document_file_not_found(self):
        """Test processing non-existent file."""
        result = self.processor.process_document('nonexistent.xlsx')
        
        assert result.processing_status == ProcessingStatus.FAILED
        assert result.content == ""
        assert "nonexistent.xlsx" in result.file_path
    
    @patch('openpyxl.load_workbook')
    def test_process_document_success(self, mock_load_workbook):
        """Test successful document processing."""
        # Create a temporary Excel file for testing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            # Create a real Excel file for testing
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'TestSheet'
            ws['A1'] = 'Name'
            ws['B1'] = 'Value'
            ws['A2'] = 'Item1'
            ws['B2'] = 100
            wb.save(tmp.name)
            wb.close()
            tmp_path = tmp.name
        
        try:
            # Mock the workbook loading
            mock_workbook = Mock()
            mock_workbook.sheetnames = ['TestSheet']
            mock_workbook.close = Mock()
            
            # Mock properties
            mock_props = Mock()
            mock_props.title = 'Test Excel'
            mock_props.creator = None
            mock_props.description = None
            mock_props.created = None
            mock_props.modified = None
            mock_workbook.properties = mock_props
            mock_workbook.active.title = 'TestSheet'
            
            # Mock worksheet
            mock_worksheet = Mock()
            mock_worksheet.max_row = 2
            mock_worksheet.max_column = 2
            mock_worksheet.iter_rows.return_value = [
                ('Name', 'Value'),
                ('Item1', 100)
            ]
            mock_worksheet._images = []  # No images
            mock_worksheet.tables = {}   # No tables
            
            mock_workbook.__getitem__.return_value = mock_worksheet
            mock_load_workbook.return_value = mock_workbook
            
            # Process document
            result = self.processor.process_document(tmp_path)
            
            # Verify results
            assert result.processing_status == ProcessingStatus.COMPLETED
            assert result.document_type == DocumentType.EXCEL
            assert 'TestSheet' in result.content
            assert 'Name\tValue' in result.content
            assert 'Item1\t100' in result.content
            assert result.metadata['total_worksheets'] == 1
            assert 'TestSheet' in result.metadata['worksheet_names']
            
            # Verify workbook was closed
            mock_workbook.close.assert_called_once()
            
        finally:
            Path(tmp_path).unlink()
    
    @patch('openpyxl.load_workbook')
    def test_process_document_with_images(self, mock_load_workbook):
        """Test document processing with images."""
        # Create a temporary Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(b'fake excel content')
            tmp_path = tmp.name
        
        try:
            # Mock workbook with images
            mock_workbook = Mock()
            mock_workbook.sheetnames = ['Sheet1']
            mock_workbook.close = Mock()
            
            # Mock properties
            mock_props = Mock()
            mock_props.title = None
            mock_props.creator = None
            mock_workbook.properties = mock_props
            mock_workbook.active.title = 'Sheet1'
            
            # Mock worksheet with images
            mock_worksheet = Mock()
            mock_worksheet.max_row = 1
            mock_worksheet.max_column = 1
            mock_worksheet.iter_rows.return_value = [('Test',)]
            
            # Mock image
            mock_image = Mock()
            mock_image._data.return_value = b'fake_image_data'
            mock_image.format = 'PNG'
            mock_image.anchor = Mock()
            mock_image.anchor._from.col = 1
            mock_image.anchor._from.row = 1
            
            mock_worksheet._images = [mock_image]
            mock_worksheet.tables = {}
            
            mock_workbook.__getitem__.return_value = mock_worksheet
            mock_load_workbook.return_value = mock_workbook
            
            # Process document
            result = self.processor.process_document(tmp_path)
            
            # Verify results
            assert result.processing_status == ProcessingStatus.COMPLETED
            assert len(result.images) == 1
            assert result.images[0].format == 'PNG'
            assert result.images[0].metadata['worksheet_name'] == 'Sheet1'
            
        finally:
            Path(tmp_path).unlink()
    
    @patch('openpyxl.load_workbook')
    def test_process_document_processing_error(self, mock_load_workbook):
        """Test document processing with error."""
        # Create a temporary Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(b'fake excel content')
            tmp_path = tmp.name
        
        try:
            # Mock openpyxl.load_workbook to raise an exception
            mock_load_workbook.side_effect = Exception("Excel parsing error")
            
            # Process document
            result = self.processor.process_document(tmp_path)
            
            # Should return failed document
            assert result.processing_status == ProcessingStatus.FAILED
            assert result.content == ""
            assert "Excel parsing error" in result.error_message
            
        finally:
            Path(tmp_path).unlink()


class TestExcelWorksheetInfo:
    """Test cases for ExcelWorksheetInfo dataclass."""
    
    def test_excel_worksheet_info_creation(self):
        """Test ExcelWorksheetInfo creation."""
        info = ExcelWorksheetInfo(
            name='TestSheet',
            max_row=100,
            max_column=10,
            cell_count=500,
            image_count=2,
            table_count=1,
            has_data=True
        )
        
        assert info.name == 'TestSheet'
        assert info.max_row == 100
        assert info.max_column == 10
        assert info.cell_count == 500
        assert info.image_count == 2
        assert info.table_count == 1
        assert info.has_data is True


class TestCellRange:
    """Test cases for CellRange dataclass."""
    
    def test_cell_range_creation(self):
        """Test CellRange creation."""
        cell_range = CellRange(
            start_row=1,
            start_col=1,
            end_row=5,
            end_col=3
        )
        
        assert cell_range.start_row == 1
        assert cell_range.start_col == 1
        assert cell_range.end_row == 5
        assert cell_range.end_col == 3
    
    def test_to_excel_range(self):
        """Test converting to Excel range notation."""
        cell_range = CellRange(
            start_row=1,
            start_col=1,
            end_row=5,
            end_col=3
        )
        
        excel_range = cell_range.to_excel_range()
        assert excel_range == 'A1:C5'
        
        # Test another range
        cell_range2 = CellRange(
            start_row=2,
            start_col=4,
            end_row=10,
            end_col=6
        )
        
        excel_range2 = cell_range2.to_excel_range()
        assert excel_range2 == 'D2:F10'


if __name__ == "__main__":
    pytest.main([__file__])