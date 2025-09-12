"""
Excel Document Processor

This module provides Excel document processing capabilities using openpyxl.
It extracts text, tables, embedded images, and metadata from Excel files for the Manufacturing RAG Agent.
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as OpenpyxlImage
from dataclasses import dataclass
import io
from PIL import Image as PILImage

from .document_processor import (
    DocumentProcessor,
    ProcessedDocument,
    DocumentType,
    ProcessingStatus,
    DocumentProcessingError,
    ExtractedImage,
    ExtractedTable,
    DocumentProcessorFactory
)


try:
    from logger.custom_logger import CustomLoggerTracker
    custom_log = CustomLoggerTracker()
    logger = custom_log.get_logger("excel_processor")

except ImportError:
    # Fallback to standard logging if custom logger not available
    logger = logging.getLogger("excel_processor")

@dataclass
class ExcelWorksheetInfo:
    """Information about an Excel worksheet."""
    name: str
    max_row: int
    max_column: int
    cell_count: int
    image_count: int
    table_count: int
    has_data: bool


@dataclass
class CellRange:
    """Represents a range of cells in Excel."""
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    
    def to_excel_range(self) -> str:
        """Convert to Excel range notation (e.g., 'A1:C5')."""
        start_col_letter = openpyxl.utils.get_column_letter(self.start_col)
        end_col_letter = openpyxl.utils.get_column_letter(self.end_col)
        return f"{start_col_letter}{self.start_row}:{end_col_letter}{self.end_row}"


class ExcelProcessor(DocumentProcessor):
    """
    Excel document processor using openpyxl.
    
    This processor extracts data from Excel worksheets, embedded images,
    and maintains proper citations with worksheet names and cell references.
    """
    
    def __init__(self, config: Dict[str, Any]):
        """
        Initialize the Excel processor.
        
        Args:
            config: Configuration dictionary containing Excel processing settings
        """
        super().__init__(config)
        self.extract_images = config.get('image_processing', True)
        self.extract_tables = config.get('table_extraction', True)
        self.min_table_rows = config.get('min_table_rows', 2)
        self.min_table_cols = config.get('min_table_cols', 2)
        self.max_empty_rows = config.get('max_empty_rows', 5)
        self.max_empty_cols = config.get('max_empty_cols', 5)
        
        logger.info(f"Excel processor initialized with image_processing={self.extract_images}, "
                   f"table_extraction={self.extract_tables}")
    
    def _get_supported_extensions(self) -> List[str]:
        """Get supported file extensions for Excel processor."""
        return ['.xlsx', '.xls', '.xlsm']
    
    def process_document(self, file_path: str) -> ProcessedDocument:
        """
        Process an Excel document and extract all content.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            ProcessedDocument with extracted content and metadata
            
        Raises:
            DocumentProcessingError: If Excel processing fails
        """
        try:
            # Validate file first
            self.validate_file(file_path)
            
            # Generate document ID
            document_id = self._generate_document_id(file_path)
            
            logger.info(f"Processing Excel document: {file_path}")
            
            # Open Excel workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            try:
                # Extract metadata
                metadata = self._extract_metadata(workbook, file_path)
                
                # Process all worksheets
                all_content = []
                all_images = []
                all_tables = []
                worksheet_info = []
                
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    logger.debug(f"Processing worksheet: {sheet_name}")
                    
                    # Extract data from worksheet
                    sheet_content = self._extract_worksheet_content(worksheet, sheet_name)
                    if sheet_content.strip():
                        all_content.append(f"[Worksheet: {sheet_name}]\n{sheet_content}")
                    
                    # Extract images if enabled
                    if self.extract_images:
                        sheet_images = self._extract_worksheet_images(worksheet, sheet_name, document_id)
                        all_images.extend(sheet_images)
                    
                    # Extract tables if enabled
                    if self.extract_tables:
                        sheet_tables = self._extract_worksheet_tables(worksheet, sheet_name)
                        all_tables.extend(sheet_tables)
                    
                    # Collect worksheet info
                    worksheet_info.append(ExcelWorksheetInfo(
                        name=sheet_name,
                        max_row=worksheet.max_row or 0,
                        max_column=worksheet.max_column or 0,
                        cell_count=self._count_non_empty_cells(worksheet),
                        image_count=len(sheet_images) if self.extract_images else 0,
                        table_count=len(sheet_tables) if self.extract_tables else 0,
                        has_data=bool(sheet_content.strip())
                    ))
                
                # Combine all content
                full_content = "\n\n".join(all_content)
                
                # Update metadata with processing info
                metadata.update({
                    'total_worksheets': len(workbook.sheetnames),
                    'worksheet_names': workbook.sheetnames,
                    'total_images': len(all_images),
                    'total_tables': len(all_tables),
                    'total_content_length': len(full_content),
                    'worksheet_info': [
                        {
                            'name': info.name,
                            'max_row': info.max_row,
                            'max_column': info.max_column,
                            'cell_count': info.cell_count,
                            'image_count': info.image_count,
                            'table_count': info.table_count,
                            'has_data': info.has_data
                        }
                        for info in worksheet_info
                    ]
                })
                
                # Create processed document
                processed_doc = ProcessedDocument(
                    document_id=document_id,
                    filename=Path(file_path).name,
                    file_path=file_path,
                    document_type=DocumentType.EXCEL,
                    content=full_content,
                    metadata=metadata,
                    images=all_images,
                    tables=all_tables,
                    processing_status=ProcessingStatus.COMPLETED
                )
                
                logger.info(f"Successfully processed Excel: {len(workbook.sheetnames)} worksheets, "
                           f"{len(all_images)} images, {len(all_tables)} tables")
                
                return processed_doc
                
            finally:
                workbook.close()
                
        except Exception as e:
            logger.error(f"Failed to process Excel {file_path}: {e}")
            
            # Create failed document
            document_id = self._generate_document_id(file_path)
            return ProcessedDocument(
                document_id=document_id,
                filename=Path(file_path).name,
                file_path=file_path,
                document_type=DocumentType.EXCEL,
                content="",
                metadata={},
                processing_status=ProcessingStatus.FAILED,
                error_message=str(e)
            )
    
    def _extract_metadata(self, workbook: openpyxl.Workbook, file_path: str) -> Dict[str, Any]:
        """
        Extract metadata from Excel workbook.
        
        Args:
            workbook: openpyxl Workbook object
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing Excel metadata
        """
        metadata = {}
        
        try:
            # Document properties
            props = workbook.properties
            
            if props.title:
                metadata['title'] = props.title
            if props.creator:
                metadata['creator'] = props.creator
            if props.description:
                metadata['description'] = props.description
            if props.subject:
                metadata['subject'] = props.subject
            if props.keywords:
                metadata['keywords'] = props.keywords
            if props.category:
                metadata['category'] = props.category
            if props.created:
                metadata['created'] = props.created.isoformat()
            if props.modified:
                metadata['modified'] = props.modified.isoformat()
            if props.lastModifiedBy:
                metadata['last_modified_by'] = props.lastModifiedBy
            
            # Workbook info
            metadata['worksheet_count'] = len(workbook.sheetnames)
            metadata['active_sheet'] = workbook.active.title if workbook.active else None
            
            # File info
            file_path_obj = Path(file_path)
            metadata['file_size'] = file_path_obj.stat().st_size
            metadata['file_extension'] = file_path_obj.suffix
            
        except Exception as e:
            logger.warning(f"Failed to extract Excel metadata: {e}")
            metadata['metadata_extraction_error'] = str(e)
        
        return metadata
    
    def _extract_worksheet_content(self, worksheet: Worksheet, sheet_name: str) -> str:
        """
        Extract content from an Excel worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            sheet_name: Name of the worksheet
            
        Returns:
            Extracted content as text
        """
        try:
            content_lines = []
            
            if not worksheet.max_row or worksheet.max_row == 1:
                return ""
            
            # Iterate through rows and columns
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                         min_col=1, max_col=worksheet.max_column,
                                         values_only=True):
                
                # Convert row values to strings, handling None values
                row_values = []
                for cell_value in row:
                    if cell_value is not None:
                        # Handle different data types
                        if isinstance(cell_value, (int, float)):
                            row_values.append(str(cell_value))
                        elif isinstance(cell_value, str):
                            row_values.append(cell_value.strip())
                        else:
                            row_values.append(str(cell_value))
                    else:
                        row_values.append("")
                
                # Skip completely empty rows
                if any(val.strip() for val in row_values if val):
                    # Join non-empty values with tabs
                    row_text = "\t".join(row_values)
                    content_lines.append(row_text)
            
            return "\n".join(content_lines)
            
        except Exception as e:
            logger.warning(f"Failed to extract content from worksheet {sheet_name}: {e}")
            return ""
    
    def _extract_worksheet_images(self, worksheet: Worksheet, sheet_name: str, document_id: str) -> List[ExtractedImage]:
        """
        Extract embedded images from an Excel worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            sheet_name: Name of the worksheet
            document_id: Document ID for image naming
            
        Returns:
            List of ExtractedImage objects
        """
        images = []
        
        try:
            # Get images from worksheet
            if hasattr(worksheet, '_images') and worksheet._images:
                for img_index, img in enumerate(worksheet._images):
                    try:
                        # Extract image data
                        image_data = self._extract_image_data(img)
                        if not image_data:
                            continue
                        
                        # Create image object
                        image_id = f"{document_id}_{sheet_name}_img{img_index}"
                        filename = f"{sheet_name}_image{img_index}.{image_data['format'].lower()}"
                        
                        # Get image position if available
                        anchor_info = {}
                        if hasattr(img, 'anchor') and img.anchor:
                            if hasattr(img.anchor, '_from'):
                                anchor_info['from_cell'] = f"{img.anchor._from.col}{img.anchor._from.row}"
                            if hasattr(img.anchor, 'to'):
                                anchor_info['to_cell'] = f"{img.anchor.to.col}{img.anchor.to.row}"
                        
                        extracted_image = ExtractedImage(
                            image_id=image_id,
                            filename=filename,
                            content=image_data['content'],
                            format=image_data['format'],
                            width=image_data.get('width'),
                            height=image_data.get('height'),
                            extraction_method="openpyxl",
                            metadata={
                                'worksheet_name': sheet_name,
                                'image_index': img_index,
                                'size_bytes': len(image_data['content']),
                                'anchor_info': anchor_info
                            }
                        )
                        
                        images.append(extracted_image)
                        
                    except Exception as e:
                        logger.warning(f"Failed to extract image {img_index} from worksheet {sheet_name}: {e}")
                        continue
        
        except Exception as e:
            logger.warning(f"Failed to extract images from worksheet {sheet_name}: {e}")
        
        return images
    
    def _extract_image_data(self, img: OpenpyxlImage) -> Optional[Dict[str, Any]]:
        """
        Extract data from an openpyxl Image object.
        
        Args:
            img: openpyxl Image object
            
        Returns:
            Dictionary with image data or None if extraction fails
        """
        try:
            # Get image data
            if hasattr(img, 'ref') and img.ref:
                # Image has reference to external file
                image_data = img.ref
            elif hasattr(img, '_data') and img._data:
                # Image data is embedded
                image_data = img._data()
            else:
                logger.warning("No image data found in image object")
                return None
            
            # Determine format
            image_format = "PNG"  # Default
            if hasattr(img, 'format') and img.format:
                image_format = img.format.upper()
            
            # Try to get dimensions using PIL
            width, height = None, None
            try:
                with io.BytesIO(image_data) as img_buffer:
                    pil_img = PILImage.open(img_buffer)
                    width, height = pil_img.size
            except Exception as e:
                logger.debug(f"Could not determine image dimensions: {e}")
            
            return {
                'content': image_data,
                'format': image_format,
                'width': width,
                'height': height
            }
            
        except Exception as e:
            logger.warning(f"Failed to extract image data: {e}")
            return None
    
    def _extract_worksheet_tables(self, worksheet: Worksheet, sheet_name: str) -> List[ExtractedTable]:
        """
        Extract tables from an Excel worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            sheet_name: Name of the worksheet
            
        Returns:
            List of ExtractedTable objects
        """
        tables = []
        
        try:
            # First, try to extract defined tables
            if hasattr(worksheet, 'tables') and worksheet.tables:
                for table_name, table in worksheet.tables.items():
                    try:
                        extracted_table = self._extract_defined_table(table, sheet_name, len(tables))
                        if extracted_table:
                            tables.append(extracted_table)
                    except Exception as e:
                        logger.warning(f"Failed to extract defined table {table_name}: {e}")
            
            # If no defined tables found, try to detect tables from data
            if not tables and self.extract_tables:
                detected_tables = self._detect_data_tables(worksheet, sheet_name)
                tables.extend(detected_tables)
        
        except Exception as e:
            logger.warning(f"Failed to extract tables from worksheet {sheet_name}: {e}")
        
        return tables
    
    def _extract_defined_table(self, table, sheet_name: str, table_index: int) -> Optional[ExtractedTable]:
        """
        Extract a defined Excel table.
        
        Args:
            table: Excel table object
            sheet_name: Name of the worksheet
            table_index: Index of the table
            
        Returns:
            ExtractedTable object or None if extraction fails
        """
        try:
            # Get table range
            table_range = table.ref
            
            # Parse range (e.g., "A1:C10")
            start_cell, end_cell = table_range.split(':')
            
            # Get table data from worksheet
            worksheet = table.parent
            table_data = []
            
            for row in worksheet[table_range]:
                row_data = []
                for cell in row:
                    value = cell.value if cell.value is not None else ""
                    row_data.append(str(value))
                table_data.append(row_data)
            
            if not table_data:
                return None
            
            # First row is typically headers
            headers = table_data[0] if table_data else []
            rows = table_data[1:] if len(table_data) > 1 else []
            
            # Create table object
            table_id = f"{sheet_name}_table{table_index}"
            
            return ExtractedTable(
                table_id=table_id,
                headers=headers,
                rows=rows,
                worksheet_name=sheet_name,
                cell_range=table_range,
                extraction_confidence=0.9,  # High confidence for defined tables
                metadata={
                    'extraction_method': 'defined_table',
                    'table_index': table_index,
                    'table_name': getattr(table, 'name', '')
                }
            )
        
        except Exception as e:
            logger.warning(f"Failed to extract defined table: {e}")
            return None
    
    def _detect_data_tables(self, worksheet: Worksheet, sheet_name: str) -> List[ExtractedTable]:
        """
        Detect tables from worksheet data patterns.
        
        Args:
            worksheet: openpyxl Worksheet object
            sheet_name: Name of the worksheet
            
        Returns:
            List of detected ExtractedTable objects
        """
        tables = []
        
        try:
            if not worksheet.max_row or worksheet.max_row < self.min_table_rows:
                return tables
            
            # Simple table detection: look for contiguous data blocks
            data_blocks = self._find_data_blocks(worksheet)
            
            for block_index, data_block in enumerate(data_blocks):
                if len(data_block) >= self.min_table_rows and len(data_block[0]) >= self.min_table_cols:
                    # Create table from data block
                    headers = data_block[0]
                    rows = data_block[1:]
                    
                    # Calculate cell range
                    start_row = 1  # This is simplified - in reality would need to track actual positions
                    end_row = start_row + len(data_block) - 1
                    start_col = 1
                    end_col = len(headers)
                    
                    cell_range = CellRange(start_row, start_col, end_row, end_col).to_excel_range()
                    
                    table_id = f"{sheet_name}_detected_table{block_index}"
                    
                    table = ExtractedTable(
                        table_id=table_id,
                        headers=headers,
                        rows=rows,
                        worksheet_name=sheet_name,
                        cell_range=cell_range,
                        extraction_confidence=0.7,  # Lower confidence for detected tables
                        metadata={
                            'extraction_method': 'data_pattern_detection',
                            'table_index': block_index
                        }
                    )
                    
                    tables.append(table)
        
        except Exception as e:
            logger.warning(f"Failed to detect data tables: {e}")
        
        return tables
    
    def _find_data_blocks(self, worksheet: Worksheet) -> List[List[List[str]]]:
        """
        Find contiguous blocks of data in the worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            
        Returns:
            List of data blocks, where each block is a list of rows
        """
        data_blocks = []
        
        try:
            current_block = []
            empty_row_count = 0
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                         min_col=1, max_col=worksheet.max_column,
                                         values_only=True):
                
                # Convert row to strings
                row_values = []
                has_data = False
                
                for cell_value in row:
                    if cell_value is not None:
                        row_values.append(str(cell_value).strip())
                        if str(cell_value).strip():
                            has_data = True
                    else:
                        row_values.append("")
                
                if has_data:
                    # Reset empty row count and add to current block
                    empty_row_count = 0
                    current_block.append(row_values)
                else:
                    # Empty row
                    empty_row_count += 1
                    
                    # If we've seen too many empty rows, end the current block
                    if empty_row_count >= self.max_empty_rows and current_block:
                        if len(current_block) >= self.min_table_rows:
                            data_blocks.append(current_block)
                        current_block = []
                        empty_row_count = 0
            
            # Add final block if it exists
            if current_block and len(current_block) >= self.min_table_rows:
                data_blocks.append(current_block)
        
        except Exception as e:
            logger.warning(f"Failed to find data blocks: {e}")
        
        return data_blocks
    
    def _count_non_empty_cells(self, worksheet: Worksheet) -> int:
        """
        Count non-empty cells in a worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            
        Returns:
            Number of non-empty cells
        """
        count = 0
        
        try:
            for row in worksheet.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value is not None and str(cell_value).strip():
                        count += 1
        except Exception as e:
            logger.warning(f"Failed to count non-empty cells: {e}")
        
        return count


# Register the Excel processor
DocumentProcessorFactory.register_processor(DocumentType.EXCEL, ExcelProcessor)